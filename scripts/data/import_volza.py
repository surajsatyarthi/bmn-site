"""
import_volza.py — ENTRY-15.0 Stage 2
Reads VOLZA xlsx files and batch-upserts rows into trade_shipments.

Usage:
  python3 import_volza.py --source-dir /path/to/xlsx/files
  python3 import_volza.py --source-dir /path/to/xlsx/files --dry-run   # count only, no DB

G3 Blueprint: PROJECT_LEDGER.md lines 1316-1351 (PM APPROVED 2026-02-24)
"""

from __future__ import annotations
import os
import re
import argparse
import logging
from datetime import date, datetime

from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

BATCH_SIZE = 1000

# ── Column aliases (lower-cased, stripped) ────────────────────────────────────
# The script does case-insensitive fuzzy header matching so minor VOLZA header
# variations across file versions are handled automatically.

COMMON_ALIASES = {
    'shipment_date':    ['date', 'shipment date', 'bill of lading date', 'b/l date'],
    'hs_code':          ['hs code', 'hs-code', 'hscode', 'customs tariff code',
                         'tariff code', 'commodity code'],
    'hs_description':   ['hs description', 'product description (hs)',
                         'commodity description'],
    'product_desc':     ['product description', 'goods description', 'description of goods',
                         'product'],
    'shipper_name':     ['shipper name', 'shipper', 'exporter name', 'exporter'],
    'shipper_address':  ['shipper address', 'exporter address'],
    'shipper_city':     ['shipper city', 'exporter city'],
    'shipper_country':  ['shipper country', 'country of origin', 'origin country'],
    'consignee_name':   ['consignee name', 'consignee', 'importer name', 'importer'],
    'consignee_address':['consignee address', 'importer address'],
    'consignee_city':   ['consignee city', 'importer city'],
    'consignee_country':['consignee country', 'destination country', 'country of destination'],
    'notify_party':     ['notify party', 'notify'],
    'quantity':         ['quantity', 'qty', 'net weight', 'net weight (kg)'],
    'quantity_unit':    ['unit', 'unit of measurement', 'uom', 'quantity unit'],
    'fob_value_usd':    ['fob value usd', 'fob value (usd)', 'fob usd', 'fob value',
                         'total fob value'],
    'cif_value_usd':    ['cif value usd', 'cif value (usd)', 'cif usd', 'cif value'],
    'port_origin':      ['port of loading', 'loading port', 'port of origin', 'origin port'],
    'port_dest':        ['port of discharge', 'discharge port', 'destination port',
                         'port of destination'],
    'shipment_mode':    ['mode of transport', 'shipment mode', 'transport mode'],
    # Indian-party contact fields (export: shipper side, import: consignee side)
    'shipper_email':    ['shipper email', 'shipper e-mail', 'exporter email'],
    'shipper_phone':    ['shipper phone', 'shipper telephone', 'exporter phone'],
    'shipper_contact':  ['shipper contact person', 'shipper contact', 'exporter contact'],
    'consignee_email':  ['consignee e-mail', 'consignee email', 'importer email'],
    'consignee_phone':  ['consignee phone', 'consignee telephone', 'importer phone'],
    'consignee_contact':['contact person', 'consignee contact person', 'consignee contact'],
    'iec':              ['iec', 'iec code', 'importer exporter code'],
}


def build_header_map(header_row: tuple) -> dict:
    """Return {alias_key: col_index} from a raw header row."""
    # Invert COMMON_ALIASES: alias_text -> field_name
    alias_lookup = {}
    for field, aliases in COMMON_ALIASES.items():
        for a in aliases:
            alias_lookup[a] = field

    result = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        field = alias_lookup.get(key)
        if field and field not in result:
            result[field] = idx
    return result


def safe_str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def safe_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y', '%d %b %Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def safe_numeric(val):
    if val is None:
        return None
    try:
        return float(str(val).replace(',', '').strip())
    except (ValueError, AttributeError):
        return None


def extract_row(raw_row: tuple, col_map: dict, trade_direction: str, source_file: str) -> dict:
    """Map one xlsx row to a trade_shipments dict."""
    def get(field):
        idx = col_map.get(field)
        return raw_row[idx] if idx is not None and idx < len(raw_row) else None

    if trade_direction == 'export':
        india_name    = safe_str(get('shipper_name'))
        india_email   = safe_str(get('shipper_email'))
        india_phone   = safe_str(get('shipper_phone'))
        india_contact = safe_str(get('shipper_contact'))
    else:  # import
        india_name    = safe_str(get('consignee_name'))
        india_email   = safe_str(get('consignee_email'))
        india_phone   = safe_str(get('consignee_phone'))
        india_contact = safe_str(get('consignee_contact'))

    return {
        'shipment_date':     safe_date(get('shipment_date')),
        'hs_code':           safe_str(get('hs_code')),
        'hs_description':    safe_str(get('hs_description')),
        'product_desc':      safe_str(get('product_desc')),
        'shipper_name':      safe_str(get('shipper_name')),
        'shipper_address':   safe_str(get('shipper_address')),
        'shipper_city':      safe_str(get('shipper_city')),
        'shipper_country':   safe_str(get('shipper_country')),
        'consignee_name':    safe_str(get('consignee_name')),
        'consignee_address': safe_str(get('consignee_address')),
        'consignee_city':    safe_str(get('consignee_city')),
        'consignee_country': safe_str(get('consignee_country')),
        'notify_party':      safe_str(get('notify_party')),
        'india_party_name':  india_name,
        'india_party_email': india_email,
        'india_party_phone': india_phone,
        'india_party_contact': india_contact,
        'india_iec':         safe_str(get('iec')),
        'quantity':          safe_numeric(get('quantity')),
        'quantity_unit':     safe_str(get('quantity_unit')),
        'fob_value_usd':     safe_numeric(get('fob_value_usd')),
        'cif_value_usd':     safe_numeric(get('cif_value_usd')),
        'port_origin':       safe_str(get('port_origin')),
        'port_dest':         safe_str(get('port_dest')),
        'shipment_mode':     safe_str(get('shipment_mode')),
        'trade_direction':   trade_direction,
        'source_file':       source_file,
    }


UPSERT_SQL = """
INSERT INTO trade_shipments (
    shipment_date, hs_code, hs_description, product_desc,
    shipper_name, shipper_address, shipper_city, shipper_country,
    consignee_name, consignee_address, consignee_city, consignee_country,
    notify_party,
    india_party_name, india_party_email, india_party_phone,
    india_party_contact, india_iec,
    quantity, quantity_unit, fob_value_usd, cif_value_usd,
    port_origin, port_dest, shipment_mode,
    trade_direction, source_file
) VALUES (
    %(shipment_date)s, %(hs_code)s, %(hs_description)s, %(product_desc)s,
    %(shipper_name)s, %(shipper_address)s, %(shipper_city)s, %(shipper_country)s,
    %(consignee_name)s, %(consignee_address)s, %(consignee_city)s, %(consignee_country)s,
    %(notify_party)s,
    %(india_party_name)s, %(india_party_email)s, %(india_party_phone)s,
    %(india_party_contact)s, %(india_iec)s,
    %(quantity)s, %(quantity_unit)s, %(fob_value_usd)s, %(cif_value_usd)s,
    %(port_origin)s, %(port_dest)s, %(shipment_mode)s,
    %(trade_direction)s, %(source_file)s
)
ON CONFLICT ON CONSTRAINT trade_shipments_dedup DO NOTHING;
"""

ENSURE_CONSTRAINT_SQL = """
DO $$
BEGIN
  IF NOT EXISTS (
    SELECT 1 FROM pg_constraint
    WHERE conname = 'trade_shipments_dedup'
  ) THEN
    ALTER TABLE trade_shipments
      ADD CONSTRAINT trade_shipments_dedup
      UNIQUE (india_party_name, shipment_date, hs_code, port_dest);
  END IF;
END$$;
"""


def get_db_connection():
    import psycopg2
    raw = open(
        os.path.join(os.path.dirname(__file__), '../../bmn-site/.env.local')
    ).read()
    # Also try env var
    db_url = os.environ.get('DATABASE_URL') or re.search(r'DATABASE_URL="(.+?)"', raw).group(1)
    # Strip pgbouncer param — psycopg2 doesn't support it
    db_url = re.sub(r'\?.*', '', db_url)
    return psycopg2.connect(db_url)


def process_file(filepath: str, trade_direction: str, conn, dry_run: bool) -> int:
    filename = os.path.basename(filepath)
    wb = load_workbook(filename=filepath, read_only=True, data_only=True)
    sheet = wb.active

    rows_iter = sheet.iter_rows(values_only=True)

    # Row 1: title/period header — skip
    try:
        next(rows_iter)
    except StopIteration:
        logger.warning(f"[{filename}] Empty file.")
        return 0

    # Row 2: column headers
    try:
        header_row = next(rows_iter)
    except StopIteration:
        logger.warning(f"[{filename}] No header row.")
        return 0

    col_map = build_header_map(header_row)
    if not col_map:
        logger.error(f"[{filename}] Could not parse headers — skipping.")
        return 0

    total_inserted = 0
    batch = []
    row_num = 2  # Already consumed 2 rows

    cur = conn.cursor() if not dry_run else None

    for raw_row in rows_iter:
        row_num += 1
        if not any(raw_row):
            continue  # skip blank rows

        record = extract_row(raw_row, col_map, trade_direction, filename)

        # Must have a valid date to be worth inserting
        if record['shipment_date'] is None:
            continue

        if dry_run:
            total_inserted += 1
            continue

        batch.append(record)

        if len(batch) >= BATCH_SIZE:
            cur.executemany(UPSERT_SQL, batch)
            conn.commit()
            total_inserted += len(batch)
            logger.info(f"[{filename}] {total_inserted}/{row_num - 2} rows inserted...")
            batch.clear()

    # Flush remaining
    if batch and not dry_run:
        cur.executemany(UPSERT_SQL, batch)
        conn.commit()
        total_inserted += len(batch)

    wb.close()

    action = "parsed (dry-run)" if dry_run else "inserted"
    logger.info(f"[{filename}] DONE — {total_inserted} rows {action}")
    return total_inserted


def detect_direction(filename: str) -> str | None:
    fn = filename.lower()
    # Match 'export'/'ex' or 'import'/'im' as tokens (separated by _, space, digit, or boundary)
    if re.search(r'export|(?<![a-z])ex(?![a-z])', fn):
        return 'export'
    if re.search(r'import|(?<![a-z])im(?![a-z])', fn):
        return 'import'
    return None


def main():
    parser = argparse.ArgumentParser(
        description='Import VOLZA xlsx files into trade_shipments (ENTRY-15.0)'
    )
    parser.add_argument('--source-dir', required=True,
                        help='Directory containing VOLZA .xlsx files')
    parser.add_argument('--dry-run', action='store_true',
                        help='Parse and count rows only — no DB writes')
    parser.add_argument('--file', default=None,
                        help='Process a single file only')
    parser.add_argument('--direction', choices=['export', 'import'], default=None,
                        help='Override auto-detection: export or import. '
                             'Required when filename contains no Ex/Im keyword.')
    args = parser.parse_args()

    if not os.path.isdir(args.source_dir):
        logger.error(f"Directory not found: {args.source_dir}")
        return 1

    # Collect files
    if args.file:
        files = [args.file] if os.path.isfile(os.path.join(args.source_dir, args.file)) else []
    else:
        files = sorted(
            f for f in os.listdir(args.source_dir)
            if f.endswith('.xlsx') and not f.startswith('~')
        )

    if not files:
        logger.error("No .xlsx files found.")
        return 1

    conn = None
    if not args.dry_run:
        logger.info("Connecting to database...")
        try:
            conn = get_db_connection()
            cur = conn.cursor()
            logger.info("Ensuring dedup constraint exists...")
            cur.execute(ENSURE_CONSTRAINT_SQL)
            conn.commit()
            logger.info("Constraint OK.")
        except Exception as e:
            logger.error(f"DB connection failed: {e}")
            return 1

    grand_total = 0
    for filename in files:
        filepath = os.path.join(args.source_dir, filename)
        # --direction flag overrides auto-detection
        if args.direction:
            direction = args.direction
        else:
            direction = detect_direction(filename)
        if direction is None:
            logger.warning(f"[{filename}] Cannot detect trade direction — skipping.")
            continue

        logger.info(f"[{filename}] Starting (direction={direction}, dry_run={args.dry_run})")
        try:
            count = process_file(filepath, direction, conn, args.dry_run)
            grand_total += count
        except Exception as e:
            logger.error(f"[{filename}] FATAL: {e}")
            if conn:
                conn.rollback()
            raise  # Per IRON RULE: stop immediately, do not self-recover

    if conn:
        conn.close()

    action = "parsed (dry-run)" if args.dry_run else "inserted"
    logger.info(f"\nGrand total: {grand_total} rows {action} across {len(files)} file(s)")
    return 0


if __name__ == '__main__':
    exit(main())
