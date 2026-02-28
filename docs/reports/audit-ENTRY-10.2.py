#!/usr/bin/env python3
"""
ENTRY-10.2 Data Asset Counter-Analysis Script
Independent verification of PM's DATA_ASSET_INVENTORY.md claims.
Iron Rule: if a file fails to read, log error and skip — never fabricate counts.
"""

import os
import sys
import csv
import zipfile
import tempfile
import traceback
from pathlib import Path
from datetime import datetime

# ── Third-party imports (graceful fallback) ──────────────────────────────────
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARN] openpyxl not available — .xlsx files will not be readable")

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False
    print("[WARN] xlrd not available — .xls files will not be readable")

DB = Path("/Users/satyarthi/Desktop/Database")
RESULTS = {}  # file_path -> {rows, cols, error, sample_headers}
ERRORS = []   # (file_path, error_message)
SKIPPED = []  # files that couldn't be opened


# ── Core readers ─────────────────────────────────────────────────────────────

def read_xlsx(path, skip_rows=0):
    """Return (row_count, headers, sample_row_1). skip_rows = number of rows to skip before headers."""
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl not installed")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return 0, [], []
    header_row = rows[skip_rows] if len(rows) > skip_rows else rows[0]
    headers = [str(c).strip() if c is not None else "" for c in header_row]
    data_rows = rows[skip_rows + 1:]
    # Count non-empty rows
    count = sum(1 for r in data_rows if any(c is not None and str(c).strip() != "" for c in r))
    sample = list(data_rows[0]) if data_rows else []
    return count, headers, sample


def read_xls(path, skip_rows=0):
    """Return (row_count, headers, sample_row_1)."""
    if not HAS_XLRD:
        raise RuntimeError("xlrd not installed")
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_index(0)
    if ws.nrows == 0:
        return 0, [], []
    header_row_idx = skip_rows
    headers = [str(ws.cell_value(header_row_idx, c)).strip() for c in range(ws.ncols)]
    count = 0
    sample = []
    for r in range(header_row_idx + 1, ws.nrows):
        row = [ws.cell_value(r, c) for c in range(ws.ncols)]
        if any(str(v).strip() != "" for v in row):
            count += 1
            if not sample:
                sample = row
    return count, headers, sample


def read_csv_file(path):
    """Return (row_count, headers, sample_row_1)."""
    encodings = ["utf-8", "utf-8-sig", "latin-1", "cp1252"]
    for enc in encodings:
        try:
            with open(path, "r", encoding=enc, errors="replace") as f:
                reader = csv.reader(f)
                rows = list(reader)
            if not rows:
                return 0, [], []
            headers = rows[0]
            count = sum(1 for r in rows[1:] if any(c.strip() for c in r))
            sample = rows[1] if len(rows) > 1 else []
            return count, headers, sample
        except Exception:
            continue
    raise RuntimeError(f"Could not read CSV with any encoding: {path}")


def auto_read(path, skip_rows=0):
    """Dispatch to correct reader based on extension."""
    p = str(path).lower()
    if p.endswith(".xlsx"):
        return read_xlsx(path, skip_rows)
    elif p.endswith(".xls"):
        return read_xls(path, skip_rows)
    elif p.endswith(".csv"):
        return read_csv_file(path)
    else:
        raise RuntimeError(f"Unsupported format: {path}")


# ── Section 1: VOLZA files ────────────────────────────────────────────────────

VOLZA_FILES = {
    "India_Plastic_Im.xlsx":       {"pm_rows": 43051, "skip": 2},
    "3304_Ex_Ind.xlsx":            {"pm_rows": 32769, "skip": 2},
    "Ind_07_Ex.xlsx":              {"pm_rows": 29625, "skip": 2},
    "Metal Scrap_Product_Import.xlsx": {"pm_rows": 24478, "skip": 2},
    "USA_India39 (1).xlsx":        {"pm_rows": 9215,  "skip": 2},
    "HS07 (1).xlsx":               {"pm_rows": 7815,  "skip": 2},
    "USA_India (1).xlsx":          {"pm_rows": 4025,  "skip": 2},
    "Export Crown Decor.xlsx":     {"pm_rows": 1957,  "skip": 2},
    "07_Ex_Can_Sing.xlsx":         {"pm_rows": 1068,  "skip": 2},
}

def audit_volza():
    print("\n" + "="*60)
    print("SECTION 1: VOLZA FILES")
    print("="*60)
    volza_results = []
    volza_dir = DB  # root of Database folder

    # First: independent discovery of all xlsx/xls files in DB root
    root_excels = sorted([f for f in volza_dir.iterdir()
                          if f.is_file() and f.suffix.lower() in (".xlsx", ".xls")
                          and not f.name.startswith(".")])
    print(f"\nAll Excel files found in Database root ({len(root_excels)} files):")
    for f in root_excels:
        print(f"  {f.name}")

    print("\n--- Verifying PM's 9 listed VOLZA files ---")
    all_dates = []

    for fname, meta in VOLZA_FILES.items():
        fpath = volza_dir / fname
        record = {"file": fname, "pm_rows": meta["pm_rows"], "actual_rows": None,
                  "delta": None, "status": None, "headers": None, "error": None,
                  "date_range": None}
        if not fpath.exists():
            record["status"] = "FILE MISSING"
            record["error"] = "File does not exist"
            ERRORS.append((fname, "File does not exist"))
            volza_results.append(record)
            print(f"  {fname}: MISSING")
            continue
        try:
            rows, headers, sample = auto_read(fpath, skip_rows=meta["skip"])
            record["actual_rows"] = rows
            record["delta"] = rows - meta["pm_rows"]
            pct = abs(record["delta"]) / meta["pm_rows"] * 100 if meta["pm_rows"] else 0
            record["status"] = "CONFIRMED" if pct <= 10 else "DISCREPANCY"
            record["headers"] = headers

            # Date range check: look for a date column
            date_col = None
            for i, h in enumerate(headers):
                if "date" in h.lower():
                    date_col = i
                    break

            if date_col is not None and HAS_XLRD and fname.endswith(".xls"):
                # Check date range for xls
                wb = xlrd.open_workbook(str(fpath))
                ws = wb.sheet_by_index(0)
                dates = []
                for r in range(meta["skip"] + 1, ws.nrows):
                    v = ws.cell_value(r, date_col)
                    if v and str(v).strip():
                        dates.append(str(v).strip())
                if dates:
                    record["date_range"] = f"{dates[0]} to {dates[-1]} (first/last)"
            elif date_col is not None and HAS_OPENPYXL and fname.endswith(".xlsx"):
                wb = openpyxl.load_workbook(str(fpath), read_only=True, data_only=True)
                ws = wb.active
                all_rows = list(ws.iter_rows(values_only=True))
                wb.close()
                dates = []
                for r in all_rows[meta["skip"] + 1:meta["skip"] + 50]:
                    if len(r) > date_col and r[date_col] is not None:
                        dates.append(str(r[date_col]))
                if dates:
                    record["date_range"] = f"Sample dates: {dates[:3]}"

            print(f"  {fname}: PM={meta['pm_rows']:,} | Actual={rows:,} | Delta={record['delta']:+,} ({pct:.1f}%) | {record['status']}")
            print(f"    Headers: {headers[:10]}")
            if record["date_range"]:
                print(f"    Dates: {record['date_range']}")
        except Exception as e:
            record["status"] = "READ_ERROR"
            record["error"] = str(e)
            ERRORS.append((fname, str(e)))
            print(f"  {fname}: READ ERROR — {e}")

        volza_results.append(record)

    # Check for VOLZA files PM missed
    pm_listed = set(VOLZA_FILES.keys())
    pm_listed.update({"Frozen Food.xlsx", "Frozen Vegetables.xlsx", "Vegetables.xlsx"})
    missed = [f.name for f in root_excels if f.name not in pm_listed]
    print(f"\nVOLZA files in root NOT in PM's list: {missed if missed else 'None'}")

    return volza_results


# ── Section 2: Santander sample ──────────────────────────────────────────────

def audit_santander_sample():
    print("\n" + "="*60)
    print("SECTION 2: SANTANDER REACH BUSINESS COUNTERPART")
    print("="*60)
    san_dir = DB / "Santander" / "Reach Business counterpart"

    if not san_dir.exists():
        print(f"  ERROR: Directory not found: {san_dir}")
        return {}

    all_files = sorted([f for f in san_dir.iterdir()
                        if f.is_file() and f.suffix.lower() in (".xls", ".xlsx")])
    print(f"\nTotal Santander files found: {len(all_files)}")
    print(f"PM claimed: 194 files")

    # Sample 20 spread across chapters
    sample_names = [
        "Ch_1.xlsx", "Ch_1(1).xlsx", "Ch_10.xls", "Ch_25.xls", "Ch_39_China.xls",
        "Ch_39_USA_Importers.xls", "Ch_50.xls", "Ch_74_Exporter.xls", "Ch_74_Importer.xls",
        "Ch_84_Exporter.xls", "Ch_84_Importer.xls", "Ch_95.xls", "Ch_2.xls", "Ch_3.xls",
        "Ch_27.xls", "Ch_39AtoF_exceptchina.xls", "Ch_52.xls", "Ch_61_Exporter.xls",
        "Ch_72_Exporter.xls", "Ch_85_Exporter.xls"
    ]

    sample_results = []
    total_sample_rows = 0
    schema_consistent = True
    first_headers = None

    for fname in sample_names:
        fpath = san_dir / fname
        if not fpath.exists():
            # Try approximate match
            matches = [f for f in all_files if fname.lower() in f.name.lower()]
            if matches:
                fpath = matches[0]
            else:
                print(f"  {fname}: NOT FOUND (skipping)")
                sample_results.append({"file": fname, "rows": None, "error": "Not found"})
                continue

        try:
            # Santander XLS: row 1 = title, row 2 = headers, data from row 3
            rows, headers, sample = auto_read(fpath, skip_rows=1)
            total_sample_rows += rows
            if first_headers is None:
                first_headers = headers
            elif set(h.lower() for h in headers if h) != set(h.lower() for h in first_headers if h):
                schema_consistent = False
                print(f"  SCHEMA DIFF in {fpath.name}: {headers}")

            # Check null rates for country column
            country_col = None
            for i, h in enumerate(headers):
                if "country" in h.lower():
                    country_col = i
                    break

            # Sample null count (just from sample rows since we can't re-read all)
            null_rate_note = f"country_col={country_col}"

            print(f"  {fpath.name}: {rows:,} rows | Headers: {headers[:8]} | {null_rate_note}")
            sample_results.append({"file": fpath.name, "rows": rows, "headers": headers, "error": None})
        except Exception as e:
            print(f"  {fpath.name}: READ ERROR — {e}")
            sample_results.append({"file": fpath.name, "rows": None, "error": str(e)})

    if sample_results:
        sampled_count = len([r for r in sample_results if r.get("rows") is not None])
        if sampled_count > 0:
            avg_rows = total_sample_rows / sampled_count
            extrapolated = int(avg_rows * len(all_files))
            print(f"\nSample: {sampled_count} files, {total_sample_rows:,} total rows, avg {avg_rows:,.0f}/file")
            print(f"Extrapolated total ({len(all_files)} files × {avg_rows:,.0f}): {extrapolated:,} rows")
            print(f"PM claimed: ~2M+ rows")
            print(f"Schema consistent across sample: {schema_consistent}")
            print(f"First file headers: {first_headers}")

    return {"total_files": len(all_files), "sample": sample_results,
            "total_sample_rows": total_sample_rows, "schema_consistent": schema_consistent,
            "first_headers": first_headers}


# ── Section 3: EPC files ─────────────────────────────────────────────────────

EPC_FILES = {
    "List of members EEPC.xlsx": {"pm_rows": 30666, "skip": 0},
    "SilkEPC.xlsx": {"pm_rows": 12787, "skip": 0},
    "spiceboard.csv": {"pm_rows": 12317, "skip": 0},
    "aepc.csv": {"pm_rows": 8150, "skip": 0},
    "epch.csv": {"pm_rows": 9020, "skip": 0},
    "pharmaexcilNoemailnophone.xlsx": {"pm_rows": 4766, "skip": 0},
    "SEZEPC.xlsx": {"pm_rows": 4659, "skip": 0},
    "CoirEPC.xlsx": {"pm_rows": 1864, "skip": 0},
    "Shefexil.xlsx": {"pm_rows": 883, "skip": 0},
    "Tobacoepcsheet.xlsx": {"pm_rows": 721, "skip": 0},
    "juteepc.xlsx": {"pm_rows": 392, "skip": 0},
    "Cashew.xlsx": {"pm_rows": 193, "skip": 0},
}

def audit_epc():
    print("\n" + "="*60)
    print("SECTION 3: EPC FILES")
    print("="*60)
    epc_dir = DB / "db"
    epc_results = []

    # Independent discovery of all files in db/
    all_db_files = sorted([f for f in epc_dir.iterdir()
                           if f.is_file() and not f.name.startswith(".")])
    print(f"\nAll files in Database/db/ root ({len(all_db_files)} files):")
    for f in all_db_files:
        print(f"  {f.name}")

    print("\n--- Verifying PM's EPC file list ---")
    for fname, meta in EPC_FILES.items():
        # Search in db/ root and subdirectories
        candidates = list(epc_dir.rglob(fname))
        if not candidates:
            print(f"  {fname}: NOT FOUND in db/")
            epc_results.append({"file": fname, "pm_rows": meta["pm_rows"], "actual_rows": None,
                                "status": "FILE MISSING", "error": "Not found"})
            continue

        fpath = candidates[0]
        try:
            rows, headers, sample = auto_read(fpath, skip_rows=meta["skip"])
            delta = rows - meta["pm_rows"]
            pct = abs(delta) / meta["pm_rows"] * 100 if meta["pm_rows"] else 0
            status = "CONFIRMED" if pct <= 10 else "DISCREPANCY"

            # Check email column
            has_email = any("email" in h.lower() for h in headers)
            has_phone = any(h.lower() in ("phone", "mobile", "telephone", "contact") for h in headers)

            print(f"  {fname}: PM={meta['pm_rows']:,} | Actual={rows:,} | Delta={delta:+,} ({pct:.1f}%) | {status}")
            print(f"    Email col: {has_email} | Phone col: {has_phone} | Headers: {headers[:8]}")
            epc_results.append({"file": fname, "pm_rows": meta["pm_rows"], "actual_rows": rows,
                                "delta": delta, "status": status, "has_email": has_email,
                                "has_phone": has_phone, "headers": headers})
        except Exception as e:
            print(f"  {fname}: READ ERROR — {e}")
            epc_results.append({"file": fname, "pm_rows": meta["pm_rows"], "actual_rows": None,
                                "status": "READ_ERROR", "error": str(e)})

    return epc_results


# ── Section 4: Embassy ZIPs ──────────────────────────────────────────────────

def audit_embassy():
    print("\n" + "="*60)
    print("SECTION 4: EMBASSY FOREIGN BUYER ZIPs")
    print("="*60)
    embassy_dir = DB / "db" / "embassy"

    if not embassy_dir.exists():
        print(f"  ERROR: Directory not found: {embassy_dir}")
        return {}

    all_files = sorted([f for f in embassy_dir.iterdir() if f.is_file() and not f.name.startswith(".")])
    zips = [f for f in all_files if f.suffix.lower() == ".zip"]
    non_zips = [f for f in all_files if f.suffix.lower() != ".zip"]

    print(f"\nFiles in embassy/: {len(all_files)} total | {len(zips)} ZIPs | {len(non_zips)} non-ZIPs")
    print("All files:")
    for f in all_files:
        print(f"  {f.name} ({f.stat().st_size:,} bytes)")

    # Non-ZIP xlsx files
    print("\n--- Non-ZIP XLSX files in embassy/ ---")
    nonzip_results = []
    for f in non_zips:
        if f.suffix.lower() in (".xlsx", ".xls"):
            try:
                rows, headers, sample = auto_read(f)
                has_email = any("email" in h.lower() or "e-mail" in h.lower() for h in headers)
                print(f"  {f.name}: {rows:,} rows | Email: {has_email} | Headers: {headers[:8]}")
                nonzip_results.append({"file": f.name, "rows": rows, "has_email": has_email, "headers": headers})
            except Exception as e:
                print(f"  {f.name}: READ ERROR — {e}")
                nonzip_results.append({"file": f.name, "rows": None, "error": str(e)})

    # Sample 5 ZIPs
    print(f"\n--- Sampling ZIPs (total {len(zips)}) ---")
    zip_results = []
    selected_zips = zips[:min(5, len(zips))]
    total_zip_rows = 0

    with tempfile.TemporaryDirectory() as tmpdir:
        for zf in selected_zips:
            zip_row_count = 0
            zip_has_email = False
            inner_files = []
            try:
                with zipfile.ZipFile(zf, "r") as z:
                    z.extractall(tmpdir)
                    inner_files = z.namelist()
            except Exception as e:
                print(f"  {zf.name}: ZIP ERROR — {e}")
                zip_results.append({"zip": zf.name, "error": str(e)})
                continue

            excel_inner = [f for f in inner_files if f.lower().endswith((".xlsx", ".xls"))]
            print(f"\n  ZIP: {zf.name} ({len(inner_files)} files, {len(excel_inner)} Excel)")

            for inner in excel_inner[:5]:  # max 5 per ZIP
                inner_path = Path(tmpdir) / inner
                if not inner_path.exists():
                    continue
                try:
                    rows, headers, sample = auto_read(inner_path)
                    has_email = any("email" in h.lower() or "e-mail" in h.lower() for h in headers)
                    if has_email:
                        zip_has_email = True
                    zip_row_count += rows
                    total_zip_rows += rows
                    print(f"    {Path(inner).name}: {rows:,} rows | Email: {has_email} | Headers: {headers[:6]}")
                except Exception as e:
                    print(f"    {Path(inner).name}: ERROR — {e}")

            zip_results.append({"zip": zf.name, "inner_count": len(inner_files),
                               "excel_count": len(excel_inner), "sample_rows": zip_row_count,
                               "has_email": zip_has_email})

    print(f"\nSampled {len(selected_zips)} ZIPs → {total_zip_rows:,} rows in sample")
    print(f"PM claimed: ~30k+ usable foreign buyer records")

    return {"total_zips": len(zips), "nonzip_files": nonzip_results,
            "zip_samples": zip_results, "sample_rows": total_zip_rows}


# ── Section 5: Zaubacorp CSVs ────────────────────────────────────────────────

def audit_zaubacorp():
    print("\n" + "="*60)
    print("SECTION 5: ZAUBACORP CSVs")
    print("="*60)

    zauba_dirs = [
        DB / "db" / "FromDesktop" / "zaubascrap" / "backend" / "csvFiles",
        DB / "db" / "FromDesktop" / "zaubascrap" / "Scrap" / "servicesepc_scrapper-main 2" / "backend" / "csvFiles",
    ]
    pm_total = 1678770
    grand_total = 0
    dir_results = []

    for d in zauba_dirs:
        print(f"\nDirectory: {d}")
        if not d.exists():
            print(f"  NOT FOUND")
            dir_results.append({"dir": str(d), "exists": False})
            continue

        csvs = sorted(d.glob("*.csv"))
        print(f"  {len(csvs)} CSV files found")
        dir_total = 0
        first_headers = None
        headers_consistent = True

        for csv_path in csvs:
            try:
                rows, headers, sample = read_csv_file(csv_path)
                dir_total += rows
                if first_headers is None:
                    first_headers = headers
                elif headers != first_headers:
                    headers_consistent = False
                print(f"  {csv_path.name}: {rows:,} rows | headers: {headers[:6]}")
            except Exception as e:
                print(f"  {csv_path.name}: ERROR — {e}")

        grand_total += dir_total
        print(f"  Dir total: {dir_total:,} rows | Schema consistent: {headers_consistent}")
        print(f"  Headers: {first_headers[:8] if first_headers else 'N/A'}")
        dir_results.append({"dir": str(d), "exists": True, "total_rows": dir_total,
                            "schema_consistent": headers_consistent, "first_headers": first_headers})

    # Are the two directories duplicates?
    print(f"\nTotal Zaubacorp rows: {grand_total:,}")
    print(f"PM claimed: ~1,678,770 rows")
    delta = grand_total - pm_total
    print(f"Delta: {delta:+,}")

    return {"pm_total": pm_total, "actual_total": grand_total,
            "delta": delta, "dirs": dir_results}


# ── Section 6: ODOP / District Export Statistics ─────────────────────────────

def audit_odop():
    print("\n" + "="*60)
    print("SECTION 6: ODOP / DISTRICT EXPORT STATISTICS")
    print("="*60)
    odop_dir = DB / "db" / "exportPatnaodop"

    if not odop_dir.exists():
        print(f"  Not found: {odop_dir}")
        return {}

    odop_files = {
        "IndiaDist21-22.xls": 219,
        "IndiaDist22-23.xls": 219,
        "IndiaDist23-24.xls": 223,
        "IndiaDistToCountry21-22.xls": 2658,
        "IndiaDistToCountry22-23.xls": 2659,
        "IndiaDistToCountry23-24.xls": 2921,
    }
    results = []
    for fname, pm_rows in odop_files.items():
        fpath = odop_dir / fname
        if not fpath.exists():
            print(f"  {fname}: MISSING")
            results.append({"file": fname, "pm_rows": pm_rows, "status": "MISSING"})
            continue
        try:
            rows, headers, sample = auto_read(fpath)
            delta = rows - pm_rows
            pct = abs(delta) / pm_rows * 100 if pm_rows else 0
            status = "CONFIRMED" if pct <= 10 else "DISCREPANCY"
            print(f"  {fname}: PM={pm_rows:,} | Actual={rows:,} | {status}")
            results.append({"file": fname, "pm_rows": pm_rows, "actual_rows": rows, "status": status})
        except Exception as e:
            print(f"  {fname}: ERROR — {e}")
            results.append({"file": fname, "pm_rows": pm_rows, "status": "READ_ERROR", "error": str(e)})
    return results


# ── Section 7: Files PM may have missed ──────────────────────────────────────

def audit_missed_files():
    print("\n" + "="*60)
    print("SECTION 7: FILES PM MAY HAVE MISSED")
    print("="*60)
    # Files in db/ root not mentioned by PM
    epc_dir = DB / "db"
    known = {
        "List of members EEPC.xlsx", "SilkEPC.xlsx", "spiceboard.csv", "aepc.csv",
        "epch.csv", "pharmaexcilNoemailnophone.xlsx", "SEZEPC.xlsx", "CoirEPC.xlsx",
        "Shefexil.xlsx", "Tobacoepcsheet.xlsx", "juteepc.xlsx", "Cashew.xlsx",
        "Mailrelay.csv", "eicelwhatsappmarketing.csv", "exportercolelctive.xlsx",
        "Buyer Seller Meet Registration Report.xlsx", "GISMPExporters.xlsx",
        "FIEO Details.xlsx", "Star export house 11.12.2022.xlsx", "UPexporter.xlsx",
        "9,000+ Investors.xlsx", "StockmarketMumbaiInvestors.xlsx", "sepcdb.xls",
        "spiceboard2ndstage.csv", "rktemail.txt", "top100teaexporters.pdf",
        "indianassociationchambers.xlsx", "jb.xls", "SEZEPC.xlsx",
        "exporters eng punjab.xlsx", "juteepc.xlsx",
        "Foreign Diplomates & Delegation GIS-23.xlsx",
        "MSME Participant Registration.xlsx",
    }
    all_files = [f for f in epc_dir.rglob("*")
                 if f.is_file() and not f.name.startswith(".")
                 and "scribedb" not in str(f)
                 and "embassy" not in str(f)
                 and "zaubascrap" not in str(f)
                 and "exportPatnaodop" not in str(f)
                 and "FromDesktop" not in str(f)]

    unknown = [f for f in all_files if f.name not in known]
    print(f"\nFiles in db/ (excl. known/scribedb/embassy/zauba/odop): {len(all_files)} total")
    print(f"Files not in PM's known list: {len(unknown)}")
    for f in unknown:
        size = f.stat().st_size
        print(f"  {f.relative_to(DB)}: {size:,} bytes")
    return [{"file": str(f.relative_to(DB)), "size": f.stat().st_size} for f in unknown]


# ── MAIN ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print(f"ENTRY-10.2 Data Asset Counter-Analysis")
    print(f"Run at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Database root: {DB}")
    print(f"openpyxl: {HAS_OPENPYXL} | xlrd: {HAS_XLRD}")

    results = {}
    results["volza"] = audit_volza()
    results["santander"] = audit_santander_sample()
    results["epc"] = audit_epc()
    results["embassy"] = audit_embassy()
    results["zaubacorp"] = audit_zaubacorp()
    results["odop"] = audit_odop()
    results["missed"] = audit_missed_files()

    print("\n" + "="*60)
    print("AUDIT COMPLETE")
    print(f"Total errors: {len(ERRORS)}")
    for e in ERRORS:
        print(f"  ERROR: {e[0]} — {e[1]}")
    print("="*60)

    # Dump compact JSON for report generation
    import json
    output_path = Path("/Users/satyarthi/Desktop/BMN/docs/reports/audit-raw-ENTRY-10.2.json")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w") as f:
        json.dump(results, f, indent=2, default=str)
    print(f"\nRaw results saved to: {output_path}")
